# -*- coding: utf-8 -*-
"""
Build inventory.xlsx — Excel/Google-Sheets version of inventory.html.

Sheets:
  - All items        (everything, with autofilter on Feature/Category/Flag)
  - Working Board
  - Sell Area
  - Finance Center
  - Meta Goal
  - Navigation & UI
  - Summary          (totals + duplicate clusters + missing-icon list)

Columns: Feature | Category | Hebrew | English | Tier/Level | Icon | Source | Flag | Notes
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUT = r"C:\Users\danza\OneDrive\Desktop\CALUDE1\funancy\inventory.xlsx"

# ---------- Inventory rows ----------------------------------------------------
# Each row: (feature, category, he, en, tier, icon, source, flag, notes)
ROWS = [
    # ---- WORKING BOARD: Producers
    ("Working Board", "Producers", "מסעדת סושי", "Sushi Restaurant (producer)", "—", "🥡", "data/producers.ts", "ok", "Spawns Sushi L1"),
    ("Working Board", "Producers", "מטבח המבורגרים", "Burger Kitchen (producer)", "—", "🍳", "data/producers.ts", "ok", "Spawns Burger L1. ProducerArea.tsx overrides with custom SVG (stove)."),
    ("Working Board", "Producers", "סטודיו אמנות", "Art Studio (producer)", "—", "🖊️", "data/producers.ts", "ok", "Spawns Art L1. ProducerArea.tsx overrides with custom SVG (paintbrush)."),
    # Sushi chain
    ("Working Board", "Sushi chain", "אורז", "Rice", "L1", "🍚", "itemChains.ts", "ok", ""),
    ("Working Board", "Sushi chain", "אונגירי", "Onigiri", "L2", "🍙", "itemChains.ts", "ok", ""),
    ("Working Board", "Sushi chain", "סושי", "Sushi", "L3", "🍣", "itemChains.ts", "ok", ""),
    ("Working Board", "Sushi chain", "מגש סושי", "Sushi Tray (Bento)", "L4", "🍱", "itemChains.ts", "ok", ""),
    ("Working Board", "Sushi chain", "דוכן ראמן", "Ramen Stand", "L5", "🏮", "itemChains.ts", "dup", "Lantern emoji also used decoratively (CSS .lantern), but those are styled, not emoji-driven."),
    ("Working Board", "Sushi chain", "מסעדת ראמן", "Ramen Restaurant", "L6", "🍜", "itemChains.ts", "ok", ""),
    ("Working Board", "Sushi chain", "מסעדת שף", "Chef Restaurant", "L7", "🏯", "itemChains.ts", "ok", "Japanese castle"),
    ("Working Board", "Sushi chain", "אימפריית סושי", "Sushi Empire", "L8 (trophy)", "🐉", "itemChains.ts", "ok", "Renders with item-trophy glow animation"),
    # Burger chain
    ("Working Board", "Burger chain", "קציצה", "Patty", "L1", "🥩", "itemChains.ts", "ok", "Cut of meat emoji"),
    ("Working Board", "Burger chain", "המבורגר", "Hamburger", "L2", "🍔", "itemChains.ts", "ok", ""),
    ("Working Board", "Burger chain", "ארוחה מהירה", "Fast Food", "L3", "🌭", "itemChains.ts", "ok", "Hot dog represents fast food"),
    ("Working Board", "Burger chain", "ארוחת קומבו", "Combo Meal", "L4", "🍟", "itemChains.ts", "ok", ""),
    ("Working Board", "Burger chain", "דוכן אוכל", "Food Stall", "L5", "🛒", "itemChains.ts", "dup", "Shopping-cart emoji as 'food stall' is semantically off; consider 🍱 stand or 🥘."),
    ("Working Board", "Burger chain", "מזנון", "Convenience / Diner", "L6", "🏪", "itemChains.ts", "ok", ""),
    ("Working Board", "Burger chain", "מסעדת גורמה", "Gourmet Restaurant", "L7", "🍽️", "itemChains.ts", "ok", ""),
    ("Working Board", "Burger chain", "רשת מסעדות", "Restaurant Chain", "L8 (trophy)", "🏆", "itemChains.ts", "dup", "Same trophy used in MetaStageBlock 'stage completed' and in GameCompleteBanner."),
    # Art chain
    ("Working Board", "Art chain", "עיפרון", "Pencil", "L1", "✏️", "itemChains.ts", "ok", ""),
    ("Working Board", "Art chain", "מכחול", "Paintbrush", "L2", "🖌️", "itemChains.ts", "ok", ""),
    ("Working Board", "Art chain", "ציור", "Painting (palette)", "L3", "🎨", "itemChains.ts", "ok", ""),
    ("Working Board", "Art chain", "ציור ממוסגר", "Framed Painting", "L4", "🖼️", "itemChains.ts", "ok", ""),
    ("Working Board", "Art chain", "תערוכה", "Exhibition", "L5", "🏛️", "itemChains.ts", "ok", ""),
    ("Working Board", "Art chain", "גלריה", "Gallery", "L6", "🎭", "itemChains.ts", "dup", "Performing-arts masks used as 'gallery' is semantically off; consider 🖼️ frame or 🏛️ variant."),
    ("Working Board", "Art chain", "מוזיאון", "Museum", "L7", "🏰", "itemChains.ts", "ok", "Castle as museum stand-in"),
    ("Working Board", "Art chain", "אימפריית אומנות", "Art Empire", "L8 (trophy)", "💎", "itemChains.ts", "ok", "Renders with item-trophy glow"),
    # Customers
    ("Working Board", "Customers (random pool)", "דני", "Danny", "—", "👨", "data/orders.ts CHARACTERS", "info", "Used by sell-request engine; OrderGallery overrides with ChefA/B SVG."),
    ("Working Board", "Customers (random pool)", "מיכל", "Michal", "—", "👩", "data/orders.ts", "info", ""),
    ("Working Board", "Customers (random pool)", "שמעון", "Shimon", "—", "👴", "data/orders.ts", "info", ""),
    ("Working Board", "Customers (random pool)", "נועה", "Noa", "—", "👸", "data/orders.ts", "info", ""),
    ("Working Board", "Customers (random pool)", "אורי", "Uri", "—", "🤴", "data/orders.ts", "info", ""),
    ("Working Board", "Customers (random pool)", "יוסי", "Yossi", "—", "👮", "data/orders.ts", "info", ""),
    ("Working Board", "Customers (random pool)", "רון", "Ron", "—", "👨\u200d🍳", "data/orders.ts", "info", ""),
    ("Working Board", "Customers (random pool)", "דנה", "Dana", "—", "👩\u200d💼", "data/orders.ts", "info", ""),
    ("Working Board", "Customers (random pool)", "אבי", "Avi", "—", "🧔", "data/orders.ts", "info", ""),
    ("Working Board", "Customers (random pool)", "רחל", "Rachel", "—", "👩\u200d🔬", "data/orders.ts", "info", ""),
    ("Working Board", "Customers (random pool)", "שרה", "Sara", "—", "👩\u200d🎨", "data/orders.ts", "info", ""),
    ("Working Board", "Customers (random pool)", "עמית", "Amit", "—", "🧑\u200d💼", "data/orders.ts", "info", ""),
    ("Working Board", "Customers (random pool)", "גיל", "Gil", "—", "👨\u200d🎓", "data/orders.ts", "info", ""),
    ("Working Board", "Customers (random pool)", "תמר", "Tamar", "—", "👩\u200d🏫", "data/orders.ts", "info", ""),
    ("Working Board", "Customers (random pool)", "יעקב", "Yaakov", "—", "🧓", "data/orders.ts", "info", ""),
    # UI hint
    ("Working Board", "UI hint", "לחץ על יצרן כדי להתחיל!", "Empty-board hint", "—", "👇", "MergeGrid.tsx", "ok", "Pulses while board is empty"),

    # ---- SELL AREA ----
    ("Sell Area", "Reward type", "מטבעות", "Coins", "—", "🪙", "data/orders.ts REWARD_EMOJIS", "dup", "Two coin glyphs in use: 🪙 (rewards, inflation) vs ⭐ (header, prices). Pick one."),
    ("Sell Area", "Reward type", "מאיץ זמן", "Time Booster", "—", "⏩", "data/orders.ts", "dup", "Time concept uses ⏩ (reward), ⏳ (button), ⌛ (header) — three glyphs."),
    ("Sell Area", "Reward type", "אנרגיה", "Energy", "—", "⚡", "data/orders.ts", "ok", "Same glyph as ResourceBar energy pill"),
    ("Sell Area", "Reward type", "תיבת הפתעה", "Mystery Box", "—", "🎁", "data/orders.ts", "ok", ""),
    ("Sell Area", "Reward type", "מאיץ מיזוג", "Merge Booster", "—", "🔮", "data/orders.ts", "ok", ""),
    ("Sell Area", "Request type tag", "פריט בודד", "Single Item", "—", "—", "OrderGallery RequestTypeTag", "missing", "Text-only pill, no icon."),
    ("Sell Area", "Request type tag", "עסקת זוג", "Duo Deal", "—", "—", "OrderGallery", "missing", "Text-only pill, no icon."),
    ("Sell Area", "Request type tag", "סט קטגוריה", "Category Set", "—", "—", "OrderGallery", "missing", "Text-only pill, no icon."),
    ("Sell Area", "Stall furniture", "קופה רושמת", "Cash Register", "—", "[SVG]", "OrderGallery CashRegister()", "ok", "Inline custom SVG, not emoji"),
    ("Sell Area", "Stall furniture", "סוכך", "Scalloped Awning", "—", "[CSS]", "globals.css .awning-scalloped", "ok", "Pure CSS"),
    ("Sell Area", "Stall furniture", "פנס", "Hanging Lantern (decorative)", "—", "[CSS]", "globals.css .lantern", "ok", "2 instances flank the awning, glow animated"),
    ("Sell Area", "Chef sprite", "שף בלונדיני", "Chef A (blonde male)", "—", "[SVG]", "OrderGallery ChefA()", "ok", "Inline SVG"),
    ("Sell Area", "Chef sprite", "שפית כובע", "Chef B (capped)", "—", "[SVG]", "OrderGallery ChefB()", "ok", "Inline SVG"),
    ("Sell Area", "UI hint", "חסרים פריטים", "'Missing items' hint", "—", "🔍", "OrderGallery", "ok", ""),

    # ---- FINANCE CENTER ----
    ("Finance Center", "Account", "חשבון בנק", "Checking Account", "—", "—", "data/accounts.ts", "missing", "AccountBreakdown shows 🏦 next to 'bank balance' overall, but each account card is text-only."),
    ("Finance Center", "Account", "פקדון בנקאי", "Bank Deposit", "—", "—", "data/accounts.ts", "missing", "No icon"),
    ("Finance Center", "Account", "קרן כספית", "Money-Market Fund", "—", "—", "data/accounts.ts", "missing", "No icon"),
    ("Finance Center", "Account", "מניה בודדת", "Single Stock", "—", "—", "data/accounts.ts", "missing", "No icon"),
    ("Finance Center", "Account", "קרן פנסיה", "Pension Fund", "—", "—", "data/accounts.ts", "missing", "No icon"),
    ("Finance Center", "Account", "קופת גמל", "Provident Fund", "—", "—", "data/accounts.ts", "missing", "No icon"),
    ("Finance Center", "Dashboard", "יתרת חשבון", "Bank Balance label", "—", "🏦", "AccountBreakdown.tsx", "ok", ""),
    ("Finance Center", "Dashboard", "השקעות", "Investments label", "—", "📈", "AccountBreakdown.tsx", "ok", ""),
    ("Finance Center", "Dashboard", "מוגן מאינפלציה", "'Protected from inflation'", "—", "🛡️", "AccountBreakdown.tsx + InflationNotification", "ok", ""),
    ("Finance Center", "Dashboard", "מאבד ערך לאינפלציה", "'Losing value' warning", "—", "⚠️", "AccountBreakdown.tsx", "ok", ""),
    ("Finance Center", "Inflation event", "אינפלציה", "Inflation hit", "—", "📉", "InflationNotification.tsx", "ok", "Shakes via .inflation-icon animation"),
    ("Finance Center", "Inflation event", "מטבעות מתעופפים", "Floating coins (lost)", "—", "🪙", "InflationNotification.tsx (×3 anims)", "dup", "Same glyph as Coins reward — see Sell Area."),
    ("Finance Center", "Inflation event", "הכסף שלך מוגן", "'Money protected' sparkle", "—", "✨", "InflationNotification.tsx", "ok", ""),
    ("Finance Center", "Inflation tip", "הפקד כסף בפקדון", "'Deposit tip'", "—", "💡", "InflationNotification.tsx", "ok", ""),
    ("Finance Center", "Time Speeder", "שנה", "Game year label", "—", "📅", "TimeSpeederButton.tsx", "ok", ""),
    ("Finance Center", "Time Speeder", "מאיץ זמן", "Time Speeder button", "—", "⏳", "TimeSpeederButton.tsx", "dup", "Resource bar uses ⌛, reward uses ⏩, this uses ⏳ — three glyphs."),
    ("Finance Center", "Time Speeder", "עברה שנה!", "'A year passed!'", "—", "📅", "TimeSpeederButton.tsx", "dup", "Same calendar reused"),
    ("Finance Center", "Mission", "פתח פיקדון ראשון", "Open First Deposit (Stage 1)", "—", "—", "data/missions.ts", "missing", "No icon defined per mission"),
    ("Finance Center", "Mission", "השתמש במאיץ זמן", "Use Time Speeder (Stage 2)", "—", "—", "data/missions.ts", "missing", "No icon"),
    ("Finance Center", "Mission", "הרוויח 5 מטבעות", "Earn Profit (Stage 3)", "—", "—", "data/missions.ts", "missing", "No icon"),
    ("Finance Center", "Mission", "מכור מוצר בדרגה 3+", "Sell Tier 3+ (Stage 4)", "—", "—", "data/missions.ts", "missing", "No icon"),
    ("Finance Center", "Mission", "פזר את ההשקעות", "Diversify (Stage 5)", "—", "—", "data/missions.ts", "missing", "No icon"),
    ("Finance Center", "UI indicator", "רווח", "Profit arrow", "—", "↑+", "AccountCard.tsx", "ok", "Used as text glyph"),
    ("Finance Center", "UI indicator", "הפסד", "Loss arrow", "—", "↓", "AccountCard.tsx", "ok", ""),
    ("Finance Center", "UI indicator", "נעול", "Locked pension", "—", "🔒", "AccountCardExpanded.tsx", "dup", "Same lock used for Meta Goal locked stages and locked items"),
    ("Finance Center", "UI control", "סגור", "Close button", "—", "✕", "AccountCardExpanded.tsx", "ok", ""),
    ("Finance Center", "UI control", "הצג השקעות", "Expand triangle", "—", "▼", "AccountBreakdown.tsx", "ok", ""),

    # ---- META GOAL ----
    ("Meta Goal", "Stage 1 header", "טכנולוגיה אישית", "Stage 1 — Personal Tech", "—", "📱", "metaGoals.ts createInitialMetaStages", "dup", "Same emoji as the Stage-1 Phone L1 item below."),
    ("Meta Goal", "Stage 1 / Phone", "טלפון מקשים", "Basic Phone", "L1 (base)", "📞", "metaGoals.ts STAGE_1_ITEMS", "ok", ""),
    ("Meta Goal", "Stage 1 / Phone", "סמארטפון", "Smartphone", "L2 (upgrade)", "📱", "metaGoals.ts", "dup", "Reused as Stage-1 stage emoji"),
    ("Meta Goal", "Stage 1 / Watch", "שעון רגיל", "Basic Watch", "L1 (base)", "🕐", "metaGoals.ts", "ok", "Clock face emoji"),
    ("Meta Goal", "Stage 1 / Watch", "שעון חכם", "Smart Watch", "L2 (upgrade)", "⌚", "metaGoals.ts", "ok", ""),
    ("Meta Goal", "Stage 1 / Earbuds", "אוזניות חוטיות", "Wired Earbuds", "L1 (base)", "🎵", "metaGoals.ts", "dup", "'Musical note' used as wired earbuds — semantically off; consider 🎙️ or custom SVG."),
    ("Meta Goal", "Stage 1 / Earbuds", "אוזניות בלוטוס", "Bluetooth Earbuds", "L2 (upgrade)", "🎧", "metaGoals.ts", "ok", ""),
    ("Meta Goal", "Stage 2 header", "עצמאות", "Stage 2 — Independence", "—", "🎓", "metaGoals.ts createInitialMetaStages", "dup", "Same emoji as the Stage-2 Studies item below."),
    ("Meta Goal", "Stage 2 / Computer", "מחשב נייח", "Desktop Computer", "L1 (base)", "🖥️", "metaGoals.ts STAGE_2_ITEMS", "ok", ""),
    ("Meta Goal", "Stage 2 / Computer", "מחשב נייד", "Laptop Computer", "L2 (upgrade)", "💻", "metaGoals.ts", "ok", ""),
    ("Meta Goal", "Stage 2 / Car", "רכב ישן", "Used Car", "L1 (base)", "🚗", "metaGoals.ts", "dup", "Same emoji as legacy LIFE_GOALS 'Used Car' (still in store for persist compat)."),
    ("Meta Goal", "Stage 2 / Car", "רכב חדש", "New Car", "L2 (upgrade)", "🚙", "metaGoals.ts", "dup", "Same emoji as legacy LIFE_GOALS 'New Car'"),
    ("Meta Goal", "Stage 2 / Studies", "לימודים", "Studies / Higher Ed", "L1 (single tier)", "🎓", "metaGoals.ts", "dup", "Reused as Stage-2 stage emoji"),
    ("Meta Goal", "Stage 3 header", "דיור ושדרוג חיים", "Stage 3 — Housing & Life Upgrade", "—", "🏡", "metaGoals.ts createInitialMetaStages", "dup", "Same emoji as Stage-3 Housing L3 below AND legacy LIFE_GOALS 'Dream Home' AND GameCompleteBanner closing line."),
    ("Meta Goal", "Stage 3 / Housing", "שכירות", "Rental", "L1 (base)", "🏚️", "metaGoals.ts STAGE_3_ITEMS", "ok", "Run-down house emoji for rental"),
    ("Meta Goal", "Stage 3 / Housing", "דירה בבניין מגורים", "Apartment in Building", "L2", "🏢", "metaGoals.ts", "ok", ""),
    ("Meta Goal", "Stage 3 / Housing", "בית צמוד קרקע", "Standalone House", "L3", "🏡", "metaGoals.ts", "dup", "Reused as Stage-3 emoji + legacy 'Dream Home' + closing banner"),
    ("Meta Goal", "Stage 3 / Vacation", "חופשה גדולה", "Big Vacation", "L1 (single tier)", "✈️", "metaGoals.ts", "ok", ""),
    ("Meta Goal", "State icon", "שלב נעול", "Locked Stage", "—", "🔒", "MetaItemCard.tsx + MetaStageBlock.tsx", "dup", "Same lock used for Finance pension and locked goals"),
    ("Meta Goal", "State icon", "הושלם", "Completed (housing/item)", "—", "✅", "MetaItemCard.tsx", "ok", ""),
    ("Meta Goal", "State icon", "שלב הושלם", "Completed Stage", "—", "🏆", "MetaStageBlock.tsx", "dup", "Same trophy as Burger L8 item AND GameCompleteBanner"),
    ("Meta Goal", "State pill", "זמין לרכישה", "'Available to buy' pill", "—", "✨", "MetaItemCard.tsx", "dup", "Same sparkle as InflationNotification 'protected'"),
    ("Meta Goal", "Header pill", "מסלול דיור", "'Housing track' header (multi-level)", "—", "🏠", "MetaItemCard.tsx MultiLevelCard", "dup", "Same emoji as legacy LIFE_GOALS 'Apartment Down-Payment'"),
    ("Meta Goal", "Completion line", "בעל בית! הושלם!", "'Homeowner!' message", "—", "🏡", "MetaItemCard.tsx", "dup", "Same as housing L3"),
    ("Meta Goal", "Game-complete banner", "מזל טוב!", "'Congratulations' header", "—", "🎉🏆🎉", "GameCompleteBanner.tsx", "dup", "Trophy reused"),
    ("Meta Goal", "Game-complete banner", "ועד בית חלומות", "Closing line house", "—", "🏡", "GameCompleteBanner.tsx", "dup", "Same as Housing L3 and Stage-3 stage emoji"),
    ("Meta Goal", "Page header", "מסלול החיים", "Goals page title", "—", "🎯", "app/goals/page.tsx", "ok", ""),
    ("Meta Goal", "Legacy LIFE_GOALS", "חופשה בחו\"ל", "Overseas Vacation", "—", "🏖️", "data/goals.ts LIFE_GOALS", "dup", "Old taxonomy; superseded by Meta Goal Stage 3 vacation. Designer can deprecate."),
    ("Meta Goal", "Legacy LIFE_GOALS", "רכב יד שנייה", "Used Car (legacy)", "—", "🚗", "data/goals.ts", "dup", "Same emoji as Meta Goal Stage-2 Used Car"),
    ("Meta Goal", "Legacy LIFE_GOALS", "רכב חדש", "New Car (legacy)", "—", "🚙", "data/goals.ts", "dup", "Same emoji as Meta Goal Stage-2 New Car"),
    ("Meta Goal", "Legacy LIFE_GOALS", "מקדמה לדירה", "Apartment Down-Payment", "—", "🏠", "data/goals.ts", "dup", "Different 'house' glyph than Meta Goal housing chain."),
    ("Meta Goal", "Legacy LIFE_GOALS", "בית החלומות", "Dream Home (legacy)", "—", "🏡", "data/goals.ts", "dup", "Same emoji as Meta Goal Stage-3 Standalone House"),

    # ---- NAVIGATION / GLOBAL UI ----
    ("Navigation / UI", "Bottom-nav tab", "מרכז פיננסי", "Finance Center tab", "—", "[SVG]", "BottomNav BankTempleIcon()", "ok", "Custom SVG: Greek-style temple with gold pediment star"),
    ("Navigation / UI", "Bottom-nav tab", "עבודה", "Working Board tab", "—", "[SVG]", "BottomNav BoardStallIcon()", "ok", "Custom SVG: market-stall with red-stripe awning"),
    ("Navigation / UI", "Bottom-nav tab", "מטרות", "Goals tab", "—", "[SVG]", "BottomNav GoalsPhotosIcon()", "ok", "Custom SVG: layered Polaroid photos (beach + car)"),
    ("Navigation / UI", "Header avatar", "אווטאר", "Player Avatar (default)", "—", "[SVG]", "ResourceBar AvatarIllustration()", "ok", "Custom SVG, no character customisation yet"),
    ("Navigation / UI", "Level badge", "רמה", "Level number badge", "—", "N°", "ResourceBar level-badge", "ok", "Numeric only, on gold gradient circle"),
    ("Navigation / UI", "Header pill", "מטבעות", "Coin balance star", "—", "⭐", "ResourceBar + everywhere", "dup", "Game uses ⭐ as coin glyph in header/prices, but reward icon for coins is 🪙. Two separate 'money' glyphs."),
    ("Navigation / UI", "Header pill", "אנרגיה", "Energy bolt", "—", "⚡", "ResourceBar + ProducerArea spawner-bolt", "ok", "Consistent everywhere"),
    ("Navigation / UI", "Header pill", "מאיצי זמן", "Time-speeder count", "—", "⌛", "ResourceBar", "dup", "Resource bar uses ⌛ but TimeSpeederButton uses ⏳ and reward uses ⏩."),
    ("Navigation / UI", "Producer SVG", "קופסת אוכל סינית", "Takeout Box (Sushi producer)", "—", "[SVG]", "ProducerArea ChainIcon('sushi')", "ok", "Inline SVG (overrides 🥡 emoji from data)"),
    ("Navigation / UI", "Producer SVG", "תנור", "Stove (Burger producer)", "—", "[SVG]", "ProducerArea ChainIcon('burger')", "ok", "Inline SVG (overrides 🍳)"),
    ("Navigation / UI", "Producer SVG", "מכחול", "Paintbrush (Art producer)", "—", "[SVG]", "ProducerArea ChainIcon('art')", "ok", "Inline SVG (overrides 🖊️)"),
    ("Navigation / UI", "Producer badge", "תג אנרגיה", "Producer corner-bolt badge", "—", "⚡", ".spawner-bolt (globals.css)", "ok", ""),
    ("Navigation / UI", "Toast notification", "הודעת מערכת", "Toast pill (slide-in)", "—", "⭐ / ⚡ / ⏩", "ui/ToastHost.tsx + toast.ts", "ok", "Uses the matching reward emoji passed to toast.success()"),
    ("Navigation / UI", "Coin badge primitive", "תג מטבע", "Beveled gold-star coin badge", "—", "[CSS+emoji]", ".ds-coin-badge with ⭐ inside", "ok", "Reusable design-system primitive"),
    ("Navigation / UI", "Progress bar", "מד התקדמות", "Gold progress bar", "—", "[CSS]", ".progress-fill-gold", "ok", "Used on missions and stage headers"),
    ("Navigation / UI", "Banner", "מאיץ מיזוג חד פעמי", "'Merge booster active' banner", "—", "🔮", "OrderGallery", "dup", "Same crystal-ball as merge_booster reward"),
    ("Navigation / UI", "Item style", "פריט גביע", "Trophy item style (any L8)", "—", "[CSS]", ".item-trophy", "ok", "Animated gold glow shared by all L8 chain items"),
    ("Navigation / UI", "Background", "רקע שוק חם", "Warm market scene gradient", "—", "[CSS]", ".bg-market-scene", "ok", "Pure CSS radial-gradient blobs"),
    ("Navigation / UI", "Background", "רקע מרכז פיננסי", "Finance Center background", "—", "[CSS]", ".bg-finance-pattern", "ok", ""),
    ("Navigation / UI", "Background", "רקע מטרות", "Goals background", "—", "[CSS]", ".bg-goals-pattern", "ok", ""),
]

HEADERS = ["Feature", "Category", "Hebrew", "English", "Tier / Level", "Icon", "Source", "Flag", "Notes"]

# Visual icon for the flag column
FLAG_GLYPH = {"ok": "✓ unique", "dup": "⚠️ duplicate", "missing": "❌ missing", "info": "• info"}

# ---------- Styling -----------------------------------------------------------
HEADER_FILL = PatternFill("solid", fgColor="FFD98708")
HEADER_FONT = Font(bold=True, color="FFFFFFFF", size=11)
ROW_OK_FILL = PatternFill("solid", fgColor="FFE8F6EA")     # green-tint
ROW_DUP_FILL = PatternFill("solid", fgColor="FFFFF4D6")    # amber-tint
ROW_MISS_FILL = PatternFill("solid", fgColor="FFFFE0DC")   # red-tint
ROW_INFO_FILL = PatternFill("solid", fgColor="FFF6F1E7")   # cream
THIN = Side(style="thin", color="FFE0CFA0")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

WRAP_LEFT = Alignment(wrap_text=True, vertical="top", horizontal="left")
WRAP_RIGHT = Alignment(wrap_text=True, vertical="top", horizontal="right", readingOrder=2)
CENTER = Alignment(wrap_text=True, vertical="center", horizontal="center")

COL_WIDTHS = [16, 22, 22, 30, 14, 12, 36, 14, 60]


def fill_for(flag):
    return {"ok": ROW_OK_FILL, "dup": ROW_DUP_FILL, "missing": ROW_MISS_FILL, "info": ROW_INFO_FILL}.get(flag, None)


def write_sheet(ws, rows, include_feature=True):
    """Write a list of inventory tuples to ws as a styled table."""
    headers = HEADERS if include_feature else HEADERS[1:]
    # Header row
    ws.append(headers)
    for col_idx, _ in enumerate(headers, start=1):
        c = ws.cell(row=1, column=col_idx)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.alignment = CENTER
        c.border = BORDER
    ws.row_dimensions[1].height = 26

    for r in rows:
        feature, cat, he, en, tier, icon, source, flag, notes = r
        flag_text = FLAG_GLYPH.get(flag, flag)
        row_values = [feature, cat, he, en, tier, icon, source, flag_text, notes] if include_feature \
            else [cat, he, en, tier, icon, source, flag_text, notes]
        ws.append(row_values)
        ws_row = ws.max_row
        fill = fill_for(flag)
        for col_idx in range(1, len(row_values) + 1):
            cell = ws.cell(row=ws_row, column=col_idx)
            cell.border = BORDER
            cell.alignment = WRAP_LEFT
            if fill is not None:
                cell.fill = fill
        # Hebrew column gets RTL alignment
        he_col = 3 if include_feature else 2
        ws.cell(row=ws_row, column=he_col).alignment = WRAP_RIGHT
        # Icon column centered + bigger font
        icon_col = 6 if include_feature else 5
        ws.cell(row=ws_row, column=icon_col).alignment = CENTER
        ws.cell(row=ws_row, column=icon_col).font = Font(size=16)

    # Column widths
    widths = COL_WIDTHS if include_feature else COL_WIDTHS[1:]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Freeze header row + autofilter
    ws.freeze_panes = "A2"
    last_col_letter = get_column_letter(len(headers))
    ws.auto_filter.ref = f"A1:{last_col_letter}{ws.max_row}"


def write_summary(ws, rows):
    ws.column_dimensions['A'].width = 40
    ws.column_dimensions['B'].width = 14
    ws.column_dimensions['C'].width = 80

    def heading(text):
        ws.append([text])
        c = ws.cell(row=ws.max_row, column=1)
        c.font = Font(bold=True, size=14, color="FFD98708")

    def kv(label, value, note=""):
        ws.append([label, value, note])
        ws.cell(row=ws.max_row, column=1).font = Font(bold=True)
        ws.cell(row=ws.max_row, column=3).alignment = WRAP_LEFT

    heading("Funancy — Visual Asset Inventory")
    ws.append([])
    kv("Total assets catalogued", len(rows))
    kv("Marked ✓ unique",  sum(1 for r in rows if r[7] == "ok"))
    kv("Marked ⚠️ duplicate", sum(1 for r in rows if r[7] == "dup"))
    kv("Marked ❌ missing icon", sum(1 for r in rows if r[7] == "missing"))
    kv("Info / customer pool",   sum(1 for r in rows if r[7] == "info"))
    ws.append([])

    heading("Per-feature counts")
    feats = {}
    for r in rows:
        feats.setdefault(r[0], {"total": 0, "dup": 0, "missing": 0})
        feats[r[0]]["total"] += 1
        if r[7] == "dup":     feats[r[0]]["dup"] += 1
        if r[7] == "missing": feats[r[0]]["missing"] += 1
    ws.append(["Feature", "Total", "⚠️ dup / ❌ missing"])
    for c in range(1, 4):
        ws.cell(row=ws.max_row, column=c).font = Font(bold=True)
        ws.cell(row=ws.max_row, column=c).fill = HEADER_FILL
        ws.cell(row=ws.max_row, column=c).font = HEADER_FONT
    for f, d in feats.items():
        ws.append([f, d["total"], f"{d['dup']} dup / {d['missing']} missing"])
    ws.append([])

    heading("Duplicate-emoji clusters (designer action items)")
    clusters = [
        ("🏡 (4 uses)",  "Stage-3 stage emoji + Housing L3 + 'Homeowner' message + GameCompleteBanner closing line + legacy LIFE_GOALS Dream Home"),
        ("🏆 (3 uses)",  "Burger L8 trophy + MetaStageBlock 'stage completed' + GameCompleteBanner"),
        ("🎓 (2 uses)",  "Stage-2 stage emoji + Studies item"),
        ("📱 (2 uses)",  "Stage-1 stage emoji + Smartphone item"),
        ("🚗 (2 uses)",  "Meta Goal Stage-2 Used Car + legacy LIFE_GOALS Used Car"),
        ("🚙 (2 uses)",  "Meta Goal Stage-2 New Car + legacy LIFE_GOALS New Car"),
        ("🏠 / 🏡",      "Different house glyphs across legacy + Meta Goal — choose one"),
        ("⌛ ⏳ ⏩",      "Three different hourglass/forward glyphs for the 'time' concept (header / button / reward)"),
        ("⭐ / 🪙",       "Two coin glyphs: ⭐ for header & prices, 🪙 for rewards & inflation animation"),
        ("🔒 (2 uses)",  "Locked pension (Finance) + locked stages/items (Meta Goal)"),
        ("✨ (2 uses)",  "InflationNotification 'protected' + MetaItemCard 'available to buy' pill"),
        ("🔮 (2 uses)",  "Merge-booster reward + 'Merge booster active' banner"),
    ]
    for label, note in clusters:
        ws.append([label, "", note])
        ws.cell(row=ws.max_row, column=1).font = Font(size=14)
        ws.cell(row=ws.max_row, column=3).alignment = WRAP_LEFT
    ws.append([])

    heading("Missing icons (designer must supply)")
    for r in rows:
        if r[7] == "missing":
            ws.append([r[3], "", f"{r[0]} — {r[1]} (source: {r[6]})"])
            ws.cell(row=ws.max_row, column=3).alignment = WRAP_LEFT


# ---------- Build workbook ----------------------------------------------------
wb = Workbook()
wb.remove(wb.active)

# Master sheet
ws_all = wb.create_sheet("All items")
write_sheet(ws_all, ROWS, include_feature=True)

# Per-feature sheets
features_in_order = ["Working Board", "Sell Area", "Finance Center", "Meta Goal", "Navigation / UI"]
sheet_names = {
    "Working Board": "Working Board",
    "Sell Area": "Sell Area",
    "Finance Center": "Finance Center",
    "Meta Goal": "Meta Goal",
    "Navigation / UI": "Navigation & UI",
}
for f in features_in_order:
    ws = wb.create_sheet(sheet_names[f])
    write_sheet(ws, [r for r in ROWS if r[0] == f], include_feature=False)

# Summary
ws_sum = wb.create_sheet("Summary", 0)  # first tab
write_summary(ws_sum, ROWS)

wb.save(OUT)
print(f"Wrote {OUT} with {len(ROWS)} rows across {len(wb.sheetnames)} sheets: {wb.sheetnames}")
